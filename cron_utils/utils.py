

import json
from django.db import connection
import requests
import xmltodict

import datetime as dt
from datetime import timedelta,datetime
from shared_config.logging import custom_log
from django.conf import settings
from shared_config.utils import export_csv
from shared_config.utils import transfer_file_to_adobe_server
from shared_config.constants import UTC_IST_DIFFERENCE_IN_HOURS
from shared_config import report_excel_sheet
from shared_config import external_config 
from shared_config import constants as api_constants
from shared_config.logging import custom_log
from shared_config.exception_constants import STATUS_TYPE, NONRETRYABLE_CODE
from shared_config.exceptions import GenericException
from shared_config import external_constants
from cron_utils.models import CPPolicies
from openpyxl import Workbook
from shared_config.utils import calculate_start_end_date,get_file_name
from shared_config.external_constants import REPORT_ACCESS_PASSWORD,TYPES_OF_PAYMENT_STATUS_FOR_REPORTS,REPORT_TYPES_FOR_GET



def get_policy_premium_details(params, request):
    """
    params : 'strUserId', 'strRefNo', 'strPolicyNo', 'strdob', 'strStatus', 'strFcode'
    """
    try:
        # NOTE: Suds client implementation was not working so had to use hardcoded input xml
        payload = external_constants.POLICY_PREMIUM_DETAILS_INPUT % (params.get('bj_user_id', ''),
                                                                     params.get('bj_ref_number', ''),
                                                                     params.get('policy_no', ''), 'NA',
                                                                     params.get('str_dob', ''))
        url = external_config.CSC_WEB_SERVICE_URL
        custom_log('info', request=request, params={'myurl': url, 'status_type': STATUS_TYPE["CP"],
                                                    'detail': 'Hitting customer portal web service.'})
        resp = requests.post(url=url, data=payload, timeout=external_constants.CUSTOMER_PORTAL_API_TIME_OUT,
                             headers={"Content-Type": "text/plain"})
        custom_log('info', request=request,
                   params={'body': {'resp_code': resp.status_code, 'response': resp.text}, 'myurl': url,
                           'status_type': STATUS_TYPE["CP"],
                           'detail': 'response from customer portal web service.'})
        if resp.status_code == external_constants.CUSTOMER_PORTAL_API_SUCCESS_CODE:
            result = resp.text
            json_response = json.loads(json.dumps(xmltodict.parse(result)))['soapenv:Envelope']['soapenv:Body'][
                'out2:GetPolicyPremiumDetails_HealthResponse']['out2:GetPolicyPremiumDetails_HealthResult']
            json_response = json.loads(json.dumps(xmltodict.parse(json_response)))
        else:
            raise GenericException(status_type=STATUS_TYPE["CP"], exception_code=NONRETRYABLE_CODE["GENERIC_FAILURE"],
                                   detail=resp.text, response_msg=api_constants.GENERIC_ERROR_MESSAGE,
                                   request=request)

        if 'Error_code' in json_response['BillJunction_Output']:
            raise GenericException(status_type=STATUS_TYPE["CP"], exception_code=NONRETRYABLE_CODE["GENERIC_FAILURE"],
                                   detail=json_response['BillJunction_Output']['Error_code'],
                                   request=request, response_msg=json_response['BillJunction_Output']['Error_code'])
        try:
            CPPolicies.objects.get_or_create(policy_no=params.get('policy_no', ''), cp_response=json_response,
                                             dob=datetime.strptime(params.get('str_dob'), '%d/%m/%Y').date())
        except Exception as e:
            raise GenericException(status_type=STATUS_TYPE["APP"], exception_code=NONRETRYABLE_CODE["BAD_REQUEST"],
                                   detail=repr(e), request=request,
                                   response_msg='')
        return json_response.get('BillJunction_Output', {}).get('Query_Output_Parameter', {})
    except GenericException as e:
        raise GenericException(status_type=STATUS_TYPE["CP"],
                               exception_code=NONRETRYABLE_CODE["GENERIC_FAILURE"],
                               detail=e.detail, request=request,
                               body=params,
                               response_msg=getattr(e, 'response_msg', api_constants.GENERIC_ERROR_MESSAGE))
    except Exception as e:
        raise GenericException(status_type=STATUS_TYPE["CP"],
                               exception_code=NONRETRYABLE_CODE["GENERIC_FAILURE"], detail=str(e), request=request)


def create_stmt(param):
    return """select
          v1.created_date,
          v1.request_count,
          v1.invalid_req,
          v2.request_sent,
          v2.success,
          v2.valid_response
        from
          (
            select
              ea1.request_count,
              ea1.created_date,
              ea2.invalid_req
            from
              (
                select
                  count(*) as request_count,
                  Date(created_at) as created_date
                from
                  public.external_apirequestlog
                where
                  service_name = '"""+param+"""'
                group by
                  created_date
              ) ea1
              left join (
                select
                  count(*) as invalid_req,
                  Date(created_at) as created_date
                from
                  public.external_apirequestlog
                where
                  status_code != 200 and service_name = '"""+param+"""'
                group by
                  created_date
              ) ea2 on ea1.created_date = ea2.created_date
          ) v1
          left join (
            select
              request_sent,
              tt1.created_date,
              success,
              valid_response
            from
              (
                select
                  t1.request_sent,
                  t1.created_date,
                  t2.success
                from
                  (
                    select
                      count(*) as request_sent,
                      Date(created_at) as created_date
                    from
                      public.external_apiexternallog
                    where
                      service_name = '"""+param+"""'
                    group by
                      created_date
                  ) t1
                  left join (
                    select
                      count(*) as success,
                      Date(created_at) as created_date
                    from
                      public.external_apiexternallog
                    where
                      status_code = 200 and service_name = '"""+param+"""'
                    group by
                      created_date
                  ) t2 on t1.created_date = t2.created_date
              ) tt1
              left join (
                select
                  count(*) as valid_response,
                  Date(created_at) as created_date
                from
                  public.external_apiexternallog
                where
                  is_valid_response = true and service_name = '"""+param+"""'
                group by
                  created_date
              ) tt2 on tt1.created_date = tt2.created_date
          ) v2 on v1.created_date = v2.created_date
        where v1.created_date >= %s and v1.created_date <= %s
        order by
          created_date asc;
        """
 
 
def agency_daily_report(param):
    """
    Execute the raw sql statement and return result as rows
    """
    with connection.cursor() as cursor:
        cursor.execute(create_stmt(param), calculate_start_end_date())
        rows = cursor.fetchall()
    return rows
 
def get_agency_monthly_report_excel_template(param):
    """
    Report excel template
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws['A1'] = "Date"
    ws['B1'] = "Total request from CRM"
    ws['C1'] = "Invalid Req"
    ws['D1'] = "Request sent to " + param
    ws['E1'] = "Success response"
    ws['F1'] = "Valid Response"
    return wb
 
 
def get_agency_monthly_report(param):
    """
    Save data in excel report and save in specified directory
    """
    wb = get_agency_monthly_report_excel_template(param)
    ws = wb.active
    data = agency_daily_report(param)
    for row, da in enumerate(data, 2):
        for col, val in enumerate(da, 1):
            ws.cell(row, col, val)
    file_name = str(get_file_name(param))
    wb.save(file_name)
    return file_name


def send_registration_data_to_adobe(params={}, request=None):
    """
    This Function is used to send user registration data to the adobe server
    Function will generate the csv report based on the params
    params can have
    {'from_date':'2018-07-01 00:00:00', 'to_date':'2018-08-01 00:00:00'} to generate_report from given date range
    Function will upload the file to the adobe server
    :return: file details
    """
    file_details = get_user_registration_data_as_csv(params, request)
    custom_log(level='info', request=request,
               params={'body': {'file_details': file_details},
                       'detail': 'created csv with data on our server and transferring to adobe sftp server'})
    # transferring file to adobe server
    transfer_file_to_adobe_server(file_details['file_address'], file_details['file_name'], request)
    custom_log(level='info', request=request,
               params={'body': {}, 'detail': 'Transferred to adobe server & send_registration_data_to_adobe'})
    
    

def get_user_registration_data_as_csv(params=None, request=None):
    """
    This function is used to generate registration data report.
    Function will generate the csv report based on the params
    params can have
    {'from_date':'2018-07-01 00:00:00', 'to_date':'2018-08-01 00:00:00'} to generate_report from given date range
    """
    if params is None:
      params={}
    custom_log('info', request=request, params={'body': {"params": params},
                                                'detail': 'adobe_cron_for_registration_data start'})
    if not params:
        params['from_date'] = dt.datetime.now() - timedelta(days=1) + timedelta(hours=UTC_IST_DIFFERENCE_IN_HOURS)
        params['to_date'] = dt.datetime.now() + timedelta(hours=UTC_IST_DIFFERENCE_IN_HOURS)
        params['from_date'] = params['from_date'].strftime(settings.INPUT_DATE_FORMAT['db_date_format'])
        params['to_date'] = params['to_date'].strftime(settings.INPUT_DATE_FORMAT['db_date_format'])
 
    file_name = params['from_date'] + "_" + params['to_date']
 
    fieldnames = ['id', 'first_name', 'last_name', 'email', 'phone', 'dob', 'gender', 'state', 'city',
                  'facebook_id', 'google_id', 'no_of_policies', 'created_at', 'login_type']
    from shared_config.utils import registered_customer_report
    data = registered_customer_report(params, request)
    params = {
        'file_name': 'registration_data_' + file_name,
        'fieldnames': fieldnames,
        'data': data,
        'sheetname': 'Registered Users'
    }
    file_details = export_csv(params, request, True)
    return file_details

def send_daily_mobile_report_sftp(request):
    """
    send daily mobile report to server. this is called from celery server.
    :return:
    """
    custom_log(level='info', request=request, params={'body': {}, 'detail': 'send_daily_report_sftp start'})
 
    yesterday = datetime.date.today() - datetime.timedelta(days=1)
    day_before_yesterday = datetime.date.today() - datetime.timedelta(days=2)
 
    params = {
        'password': REPORT_ACCESS_PASSWORD,
        'from_date': str(day_before_yesterday),
        'to_date': str(yesterday),
        'generate_new': True,
        'email': False,
        'report_types': REPORT_TYPES_FOR_GET,
        'payment_status': TYPES_OF_PAYMENT_STATUS_FOR_REPORTS['success']
    }
    custom_log(level='info', request=request,
               params={'body': {}, 'detail': 'creating of file start'})
 
    file_details = report_excel_sheet(params, request)
    custom_log(level='info', request=request,
               params={'body': {}, 'detail': 'created file and sending and transfer file start'})
    # transferring file to tebt
    try:
        transfer_file_to_adobe_server(file_details['file_address'], 'report_' + file_details['file_name'], request,
                                      settings.MOBILE_DATA_REPORTS_SERVER_DETAILS)
    except GenericException as e:
        raise GenericException(status_type=STATUS_TYPE["APP"], exception_code=NONRETRYABLE_CODE["BAD_REQUEST"],
                               detail=e.detail, response_msg=e.detail, request=request)
    except Exception as e:
        raise GenericException(status_type=STATUS_TYPE["APP"], exception_code=NONRETRYABLE_CODE["BAD_REQUEST"],
                               detail=str(repr(e)), response_msg=getattr(repr(e), 'response_msg',
                                                                         api_constants.GENERIC_ERROR_MESSAGE),
                               request=request)
    custom_log(level='info', request=request,
               params={'body': {}, 'detail': 'Transferred to adobe server & send_daily_report_sftp end'})